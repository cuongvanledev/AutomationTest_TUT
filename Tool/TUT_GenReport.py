import openpyxl 
from xml.dom import minidom
import sys
import os
import re
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, numbers
import shutil

def workSpace(module):
    res = "U:\\internal\\Module\\" + module.lower() + "\\07_UT\\01_WorkProduct_T\\workspace\\Output\\"
    return res

def adjustWidth(sheetName, pathFile):
    workbook = openpyxl.load_workbook(str(pathFile))
    workbook.active
    worksheet = workbook.get_sheet_by_name(sheetName)

    # Start changing width from column C onwards
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name# Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based) 
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        if str(column) == "A" or str(column) == "B":
            adjusted_width = 10
        else:
            adjusted_width = (max_length + 1)
        worksheet.column_dimensions[column].width = adjusted_width
    workbook.save(pathFile)
    workbook.close()

def adjustWidths(sheetNames, pathFile):
    workbook = openpyxl.load_workbook(str(pathFile))
    workbook.active
    for sheetName in sheetNames:
        if sheetName == "Cover":
            continue
        worksheet = workbook.get_sheet_by_name(sheetName)

        # Start changing width from column C onwards
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name# Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based) 
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            if str(column) == "A" or str(column) == "B":
                adjusted_width = 10
            elif str(column) == "C" and max_length > 110:
                adjusted_width = 110
            else: 
                adjusted_width = (max_length + 1) 
            worksheet.column_dimensions[column].width = adjusted_width
    workbook.save(pathFile)
    workbook.close()

# Function: summarySheet
#  Copy RH850_X2x_Msn_TUT_CR_U2A8_Beta.xlsx to workspace
#  Remove all sheet, except Cover sheet
#  Store data into Result Summary sheet
def summarySheet(dictClass, module):
    Output = workSpace(module)
    shutil.copy2("U:/internal/Module/" + module.lower() + "/07_UT/01_WorkProduct_T/result/U2A8/Beta/test_report/RH850_X2x_" + module.upper() + "_TUT_CR_U2A8_Beta.xlsx", Output) # target filename is /dst/dir/file.ext
    fileName = "RH850_X2x_" + module.upper() + "_TUT_CR_U2A8_Beta.xlsx"
    pathFile_ = workSpace(module) + fileName
    removeSheet(pathFile_)
    workbook = openpyxl.load_workbook(pathFile_)
    workbook.active
    sheet_name = "Result Summary"
    title = module.upper() + " Coverage Result"
    if sheet_name in workbook:
        #f.write("[NG] Sheet: " + sheet_name + " is existed\n")
        return
    worksheet = workbook.create_sheet(sheet_name)
    rowCurrent = 1 
    font = Font(name='Calibri',
                 size=20,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')
    font_2 = Font(name='Calibri',
             size=11,
             bold=True,
             italic=False,
             vertAlign=None,
             underline='none',
             strike=False,
             color='FF000000')
    worksheet.cell(row=rowCurrent, column=3).value  = title
    worksheet.cell(row=rowCurrent, column=3).font = font
    rowCurrent += 3
    
    #merge cells
    worksheet.merge_cells('B4:B5')
    worksheet.merge_cells('C4:C5')
    worksheet.merge_cells('D4:G4')
    worksheet.merge_cells('H4:K4')
    worksheet.merge_cells('L4:L5')
    rows = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L']
    #Color fill
    color = PatternFill(start_color='B4C6E7',
                   end_color='B4C6E7',
                   fill_type='solid')
    for r in rows:
        worksheet[r + str(4)].fill = color
        worksheet[r + str(5)].fill = color
    #border
    cell_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000')
    )

    worksheet.cell(row=rowCurrent, column=2).value  = "No"
    worksheet.cell(row=rowCurrent, column=3).value  = "Class"
    worksheet.cell(row=rowCurrent, column=4).value  = "Block Covered"
    worksheet.cell(row=rowCurrent, column=8).value  = "Line Covered"  
    worksheet.cell(row=rowCurrent, column=12).value  = "Remark"
    rowCurrent += 1
    worksheet.cell(row=rowCurrent, column=4).value  = "% Blocks Covered"
    worksheet.cell(row=rowCurrent, column=5).value  = "% Reviewed"
    worksheet.cell(row=rowCurrent, column=6).value  = "Blocks Covered"
    worksheet.cell(row=rowCurrent, column=7).value  = "Blocks Not Covered"
    worksheet.cell(row=rowCurrent, column=8).value  = "% Lines Covered"
    worksheet.cell(row=rowCurrent, column=9).value  = "%Reviewed"
    worksheet.cell(row=rowCurrent, column=10).value  = "Lines Covered"
    worksheet.cell(row=rowCurrent, column=11).value  = "Lines Not Covered"
    #worksheet.cell(row=rowCurrent, column=12).value  = "% PartiallyCover"
    #worksheet.cell(row=rowCurrent, column=13).value  = "PartiallyCover"

    for i in range(2, 13):
        worksheet.cell(row=rowCurrent - 1, column=i).border = cell_border
        worksheet.cell(row=rowCurrent, column=i).border = cell_border
        worksheet.cell(row=rowCurrent - 1, column=i).font = font_2
        worksheet.cell(row=rowCurrent, column=i).font = font_2
        worksheet.cell(row=rowCurrent, column=i).alignment = Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=rowCurrent - 1, column=i).alignment = Alignment(horizontal='center', vertical='center')

    rowCurrent += 1
    index = 1
    listClass = list(dictClass.keys())
    listClass.sort()
    for cl in listClass:
        remark = re.search("_temp", cl)
        rem = "-"
        clsName = cl

        
        if (float(dictClass[cl]['block_coverage'])) == 0:
            rem = "Just initializing Instance, no logic method to test"

        if remark:
            clsName = cl[:-5]
            rem = "TUT_" + module.upper() + "E2xCommonize(There are differences  between properties (get;set)of each devices so the % covered is different.)"

        worksheet.cell(row=rowCurrent, column=2).value  = index
        worksheet.cell(row=rowCurrent, column=3).value  = clsName
        worksheet.cell(row=rowCurrent, column=4).value  = float(dictClass[cl]['block_coverage'])* 1.0
        worksheet.cell(row=rowCurrent, column=4).number_format = numbers.FORMAT_PERCENTAGE_00
        worksheet.cell(row=rowCurrent, column=5).value  = float(dictClass[cl]['block_not_coverage'])* 1.0  
        worksheet.cell(row=rowCurrent, column=5).number_format = numbers.FORMAT_PERCENTAGE_00
        worksheet.cell(row=rowCurrent, column=6).value  = float(dictClass[cl]['blocks_covered'])
        worksheet.cell(row=rowCurrent, column=7).value  = float(dictClass[cl]['blocks_not_covered'])

        worksheet.cell(row=rowCurrent, column=8).value  = float(dictClass[cl]['line_coverage'])* 1.0
        worksheet.cell(row=rowCurrent, column=8).number_format = numbers.FORMAT_PERCENTAGE_00
        worksheet.cell(row=rowCurrent, column=9).value  = float(dictClass[cl]['line_not_coverage'])* 1.0  
        worksheet.cell(row=rowCurrent, column=9).number_format = numbers.FORMAT_PERCENTAGE_00
        worksheet.cell(row=rowCurrent, column=10).value  = float(dictClass[cl]['lines_covered'])
        worksheet.cell(row=rowCurrent, column=11).value  = float(dictClass[cl]['lines_not_covered'])
        #worksheet.cell(row=rowCurrent, column=12).value  = float(dictClass[cl]['PartiallyCover_100']) *1.0
        #worksheet.cell(row=rowCurrent, column=12).number_format = numbers.FORMAT_PERCENTAGE_00
        #worksheet.cell(row=rowCurrent, column=13).value  = float(dictClass[cl]['PartiallyCover'])
        worksheet.cell(row=rowCurrent, column=12).value  = rem

        for i in range(2, 13):
            worksheet.cell(row=rowCurrent, column=i).border = cell_border
        index += 1
        rowCurrent += 1
    #Fill remark for the same class
    amountOfRows = worksheet.max_row
    for k in range(6, amountOfRows):
        data = str(worksheet.cell(row=k, column=3).value) + "_temp"
        if (data in dictClass) and (str(worksheet.cell(row=k, column=12).value) == "-"):
            newCell = "TUT_" + module.upper() + "U2xCommonize(There are differences  between properties (get;set)of each devices so the % covered is different.)"
            worksheet.cell(row=k, column=12).value = newCell
    workbook.save(workSpace(module) + fileName)
    workbook.close()
    adjustWidth(sheet_name, workSpace(module) + fileName)
    #copyCoverSheet(module)

def getSameClass(listns, nsDict):
    clsList = dict()
    for ns in listns:
        listCL  = list(nsDict[ns].keys())
        for cl in listCL:
            if (re.search("_temp", cl)):
                if ns in clsList:
                    clsList[ns].update({cl : nsDict[ns][cl]})
                    clsList[ns].update({cl[:-5] : nsDict[ns][cl[:-5]]})
                else:
                    clsList[ns] = dict()
                    clsList[ns].update({cl : nsDict[ns][cl]})
                    clsList[ns].update({cl[:-5] : nsDict[ns][cl[:-5]]})
        
    return clsList


def classSheet(nsDict, module, clDict, commit):
    listNameSpace = list(nsDict.keys())
    fileName = "RH850_X2x_" + module.upper() + "_TUT_CR_U2A8_Beta.xlsx"
    
    workbook = openpyxl.load_workbook(str(workSpace(module) + fileName))
    workbook.active
    listSheetName = list()
    for ns in listNameSpace:
        listClass = list(nsDict[ns].keys())
        clsDict = getSameClass(listNameSpace, nsDict)
        flag = False
        for cl in listClass:
            clPar = cl
            if (ns in clsDict) and (cl in clsDict[ns]):
                if (re.search("_temp$", cl)):
                    flag = True
                    clPar = cl[:-5]
                else:
                    continue
             
            sn = ns + "." + clPar 
            sheet_name = (sn.split(".")[-3] + "." + sn.split(".")[-1])[:31]
            worksheet = workbook.create_sheet(sheet_name)
            listSheetName.append(sheet_name)
            title = clPar + " Coverage Result"
            rowCurrent = 1 
            font = Font(name='Calibri',
                         size=20,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='FF000000')
            font_2 = Font(name='Calibri',
                     size=11,
                     bold=True,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FF000000')
            worksheet.cell(row=rowCurrent, column=3).value  = title
            worksheet.cell(row=rowCurrent, column=3).font = font
            rowCurrent += 3

            #merge cells
            worksheet.merge_cells('B4:B5')
            worksheet.merge_cells('C4:C5')
            worksheet.merge_cells('D4:G4')
            worksheet.merge_cells('H4:K4')
            rows = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
            #Color fill
            color = PatternFill(start_color='B4C6E7',
                   end_color='B4C6E7',
                   fill_type='solid')

            color_coverage = PatternFill(start_color='FF0000',
                    end_color='FF0000',
                    fill_type='solid')
            for r in rows:
                worksheet[r + str(4)].fill = color
                worksheet[r + str(5)].fill = color
            #border
            cell_border = Border(left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'),
                                 top=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000')
            )

            ran = 2
            if flag == True:
                ran = 3
            for count in range(ran):
                col = 12
                if count == 0:
                    ns_class = "Class"
                else:
                    col = 14
                    ns_class = "Method"
                    worksheet.cell(row=rowCurrent, column=12).value  = "Code Review Comment"
                    worksheet.cell(row=rowCurrent, column=13).value  = "Uncovered code"
                    worksheet.merge_cells(start_row=rowCurrent, start_column=12, end_row=rowCurrent+1, end_column=12)
                    worksheet.merge_cells(start_row=rowCurrent, start_column=13, end_row=rowCurrent+1, end_column=13)
                    rows1 = ['L', 'M']
                    for r in rows1:
                        worksheet[r + str(rowCurrent)].fill = color
                        worksheet[r + str(rowCurrent + 1)].fill = color

                    #dictClass = nsDict[]
                worksheet.cell(row=rowCurrent, column=2).value  = "No"
                worksheet.cell(row=rowCurrent, column=3).value  = ns_class
                worksheet.cell(row=rowCurrent, column=4).value  = "Block Covered"
                worksheet.cell(row=rowCurrent, column=8).value  = "Line Covered"  
                rowCurrent += 1
                worksheet.cell(row=rowCurrent, column=4).value  = "% Blocks Covered"
                worksheet.cell(row=rowCurrent, column=5).value  = "% Reviewed"
                worksheet.cell(row=rowCurrent, column=6).value  = "Blocks Covered"
                worksheet.cell(row=rowCurrent, column=7).value  = "Blocks Not Covered"
                worksheet.cell(row=rowCurrent, column=8).value  = "% Lines Covered"
                worksheet.cell(row=rowCurrent, column=9).value  = "%Reviewed"
                worksheet.cell(row=rowCurrent, column=10).value  = "Lines Covered"
                worksheet.cell(row=rowCurrent, column=11).value  = "Lines Not Covered"
                #worksheet.cell(row=rowCurrent, column=12).value  = "% PartiallyCover"
                #worksheet.cell(row=rowCurrent, column=13).value  = "PartiallyCover"
                worksheet.merge_cells(start_row=rowCurrent - 1, start_column=2, end_row=rowCurrent, end_column=2)
                worksheet.merge_cells(start_row=rowCurrent - 1, start_column=3, end_row=rowCurrent, end_column=3)
                worksheet.merge_cells(start_row=rowCurrent - 1, start_column=4, end_row=rowCurrent - 1, end_column=7)
                worksheet.merge_cells(start_row=rowCurrent - 1, start_column=8, end_row=rowCurrent - 1, end_column=11)
                rows2 = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
                for r in rows2:
                    worksheet[r + str(rowCurrent - 1)].fill = color
                    worksheet[r + str(rowCurrent)].fill = color

                for i in range(2, col):
                    worksheet.cell(row=rowCurrent - 1, column=i).border = cell_border
                    worksheet.cell(row=rowCurrent, column=i).border = cell_border
                    worksheet.cell(row=rowCurrent - 1, column=i).font = font_2
                    worksheet.cell(row=rowCurrent, column=i).font = font_2
                    worksheet.cell(row=rowCurrent, column=i).alignment = Alignment(horizontal='center', vertical='center')
                    worksheet.cell(row=rowCurrent - 1, column=i).alignment = Alignment(horizontal='center', vertical='center')
                rowCurrent += 1
                index = 1

                if (flag == True) and (count == 0):
                    listCl = [ns + "." + cl, ns + "." + cl[:-5]]
                    dictClass = clDict
                elif (flag == True) and (count == 1):
                    listCl = list(nsDict[ns][cl[:-5]].keys())
                    dictClass = nsDict[ns][cl[:-5]]
                    flag = False
                elif count == 0:
                    temp = ns + "." + cl
                    listCl = [temp]
                    dictClass = clDict
                else:
                    listCl = list(nsDict[ns][cl].keys())
                    listCl.sort()
                    dictClass = nsDict[ns][cl]

                for mt in listCl:
                    remark = re.search("_temp", mt)
                    cl_ = mt
                    if remark:
                        cl_ = mt[:-5]
                    worksheet.cell(row=rowCurrent, column=2).value  = index
                    worksheet.cell(row=rowCurrent, column=3).value  = cl_
                    worksheet.cell(row=rowCurrent, column=4).value  = float(dictClass[mt]['block_coverage'])* 1.0 
                    worksheet.cell(row=rowCurrent, column=4).number_format = numbers.FORMAT_PERCENTAGE_00
                    worksheet.cell(row=rowCurrent, column=5).value  = float(dictClass[mt]['block_not_coverage'])* 1.0  
                    worksheet.cell(row=rowCurrent, column=5).number_format = numbers.FORMAT_PERCENTAGE_00
                    worksheet.cell(row=rowCurrent, column=6).value  = float(dictClass[mt]['blocks_covered'])
                    worksheet.cell(row=rowCurrent, column=7).value  = float(dictClass[mt]['blocks_not_covered'])

                    worksheet.cell(row=rowCurrent, column=8).value  = float(dictClass[mt]['line_coverage'])* 1.0
                    worksheet.cell(row=rowCurrent, column=8).number_format = numbers.FORMAT_PERCENTAGE_00
                    worksheet.cell(row=rowCurrent, column=9).value  = float(dictClass[mt]['line_not_coverage'])* 1.0  
                    worksheet.cell(row=rowCurrent, column=9).number_format = numbers.FORMAT_PERCENTAGE_00
                    worksheet.cell(row=rowCurrent, column=10).value  = float(dictClass[mt]['lines_covered'])
                    worksheet.cell(row=rowCurrent, column=11).value  = float(dictClass[mt]['lines_not_covered'])


                    if count != 0:
                        worksheet.cell(row=rowCurrent, column=13).value  = "-"
                        if (float(dictClass[mt]['block_coverage'])) == 0:
                            if mt[:-2] == clPar:
                                worksheet.cell(row=rowCurrent, column=12).value = "No test for Constructor (no code logic)"
                            else:
                                worksheet.cell(row=rowCurrent, column=12).value = "No test for Property variables (no code logic)"
                        elif (float(dictClass[mt]['block_coverage'])) > 0 and (float(dictClass[mt]['block_coverage'])) < 1 and commit.lower() == "no":
                            worksheet.cell(row=rowCurrent, column=12).value = "-"
                            worksheet.cell(row=rowCurrent, column=12).fill  = color_coverage
                        else:
                            worksheet.cell(row=rowCurrent, column=12).value = "-"

                    for i in range(2, col):
                        worksheet.cell(row=rowCurrent, column=i).border = cell_border
                    index += 1
                    rowCurrent += 1
                rowCurrent += 3
    
    workbook.save(workSpace(module) + fileName)
    workbook.close()
    listSheets = template(module)
    adjustWidths(listSheets, workSpace(module) + fileName)
    hyperLink(workSpace(module) + fileName)
    return

def genLogFile(pathFile, Msn, Device):
    
    xmldoc = minidom.parse(pathFile)
    testRun = xmldoc.getElementsByTagName('TestRun')
    ResultSummary = testRun[0].getElementsByTagName('ResultSummary')
    Counters = ResultSummary[0].getElementsByTagName('Counters')
    pathLog = os.path.dirname(os.path.abspath(pathFile))
    logFile = "RH850_X2x_" + Msn.upper() + "_" + Device + "_TUT_TestLog.log"
    if Device == "None":
        Device = ""
        logFile = "RH850_X2x_" + Msn.upper() + "_TUT_TestLog.log"
    #Creat file log
    f = open(pathLog + "\\"  +  "\\" + logFile, "a")
    #
    f.write("Group Name: "+ "TUT_" + Msn + Device + "Commonize" + "\n")
    f.write("Group By: Hierarchy" + "\n")
    f.write("Group Full Name: "+ "TUT_" + Msn + Device + "Commonize" + "\n")
    f.write("Total TestCases: " + Counters[0].getAttribute('total') + "\n")
    f.write("Executed: " + Counters[0].getAttribute('executed') + " test(s)" + "\n")
    f.write("Not Executed: " + Counters[0].getAttribute('notExecuted') + " test(s)" + "\n")
    f.write("Passed: " + Counters[0].getAttribute('passed') + " test(s)" + "\n")
    f.write("Failed: " + Counters[0].getAttribute('failed') + " test(s)" + "\n")
    f.write("\n")
    #
    index = 1
    for testCase in testRun:
        allTCs = testCase.getElementsByTagName('Results')
        UnitTestResult = allTCs[0].getElementsByTagName('UnitTestResult')
        for tc in UnitTestResult:
            resultTC = tc.getAttribute('outcome')
            reason = ""
            if resultTC == "Failed":
                reason = tc.getElementsByTagName('Output')[0].getElementsByTagName('ErrorInfo')[0].getElementsByTagName('Message')[0].firstChild.nodeValue
            f.write("Result" + str(index) + " Name: "+ tc.getAttribute("testName") + "\n")
            f.write("Result" + str(index) + " Outcome: " + tc.getAttribute('outcome') + "\n")
            f.write("Result" + str(index) + " Duration: " + tc.getAttribute('duration') + "\n")
            f.write("Result" + str(index) + " StartTime: " + tc.getAttribute('startTime') + "\n")
            f.write("Result" + str(index) + " EndTime: " + tc.getAttribute('endTime') + "\n")
            f.write("Result" + str(index) + " Message: " + reason + "\n")
            f.write("Result" + str(index) + " StandardOutput:" + "\n")
            f.write("Result" + str(index) + " StandardError:" + "\n")
            f.write("\n")
            index = index + 1

# Remove all sheet, except Cover sheet
def removeSheet(pathFile):
    workbook = openpyxl.load_workbook(pathFile)
    workbook.active
    sheetNames = workbook.get_sheet_names()
    for sheet in sheetNames:
        if sheet == "Cover":
            continue
        workbook.remove(workbook[sheet])
    workbook.save(str(pathFile)) 
    workbook.close()

def hyperLink(pathFile):
    workbook = openpyxl.load_workbook(pathFile)
    workbook.active
    ws = workbook.get_sheet_names()
    for sheetName in ws:
        worksheet = workbook.get_sheet_by_name(sheetName)
        if sheetName == "Cover":
            continue
        if sheetName != "Result Summary":
            #Sheet = worksheet.get_sheet_by_name(sheetName)
            link = "#'Result Summary'!A1"
            worksheet.cell(row=1, column=1).value = "Back"
            worksheet.cell(row=1, column=1).hyperlink = link
            worksheet.cell(row=1, column=1).font = Font(name='Calibri', size=12,color="000000FF", underline='single')
        else:
            maxrows = worksheet.max_row
            for i in range(6, maxrows + 1):
                value_ = str(worksheet.cell(row=i, column=3).value)
                
                if value_ is None:
                    continue
                else:
                    sheet_Name = value_.split(".")[-3] + "." + value_.split(".")[-1]
                    sheetHyperLink = '#' + "'" + sheet_Name[0:31] + "'"
                    link = sheetHyperLink + "!A1"
                    worksheet.cell(row = i, column=3).hyperlink = link
                    worksheet.cell(row= i, column=3).font = Font(name='Calibri', size=12,color="000000FF", underline='single')
            
    workbook.save(pathFile)
    workbook.close()

def template(module):
    fileName = "RH850_X2x_" + module.upper() + "_TUT_CR_U2A8_Beta.xlsx"
    rowCurrent = 1
    workbook = openpyxl.load_workbook(str(workSpace(module) + fileName))
    workbook.active
    worksheet = workbook.create_sheet("Template")
    font = Font(name='Calibri',
                 size=20,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')
    font_2 = Font(name='Calibri',
             size=11,
             bold=True,
             italic=False,
             vertAlign=None,
             underline='none',
             strike=False,
             color='FF000000')
    worksheet.cell(row=rowCurrent, column=3).value  = "<Class> Coverage Result"
    worksheet.cell(row=rowCurrent, column=3).font = font
    rowCurrent += 3

    #border
    cell_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000')
    )

    rows = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    #Color fill
    color = PatternFill(start_color='B4C6E7',
           end_color='B4C6E7',
           fill_type='solid')
    for count in range(2):
        col = 12

        if count == 0:
            ns_class = "Class"
        else:
            col = 14
            ns_class = "Method"
            worksheet.cell(row=rowCurrent, column=12).value  = "Code Review Comment"
            worksheet.cell(row=rowCurrent, column=13).value  = "Uncovered code"
            worksheet.merge_cells(start_row=rowCurrent, start_column=12, end_row=rowCurrent+1, end_column=12)
            worksheet.merge_cells(start_row=rowCurrent, start_column=13, end_row=rowCurrent+1, end_column=13)
            rows.append('L')
            rows.append('M')
            #dictClass = nsDict[]
        worksheet.cell(row=rowCurrent, column=2).value  = "No"
        worksheet.cell(row=rowCurrent, column=3).value  = ns_class
        worksheet.cell(row=rowCurrent, column=4).value  = "Block Covered"
        worksheet.cell(row=rowCurrent, column=8).value  = "Line Covered"  
        rowCurrent += 1
        worksheet.cell(row=rowCurrent, column=4).value  = "% Blocks Covered"
        worksheet.cell(row=rowCurrent, column=5).value  = "% Reviewed"
        worksheet.cell(row=rowCurrent, column=6).value  = "Blocks Covered"
        worksheet.cell(row=rowCurrent, column=7).value  = "Blocks Not Covered"
        worksheet.cell(row=rowCurrent, column=8).value  = "% Lines Covered"
        worksheet.cell(row=rowCurrent, column=9).value  = "%Reviewed"
        worksheet.cell(row=rowCurrent, column=10).value  = "Lines Covered"
        worksheet.cell(row=rowCurrent, column=11).value  = "Lines Not Covered"
        for r in rows:
            worksheet[r + str(rowCurrent - 1)].fill = color
            worksheet[r + str(rowCurrent)].fill = color

        worksheet.merge_cells(start_row=rowCurrent - 1, start_column=2, end_row=rowCurrent, end_column=2)
        worksheet.merge_cells(start_row=rowCurrent - 1, start_column=3, end_row=rowCurrent, end_column=3)
        worksheet.merge_cells(start_row=rowCurrent - 1, start_column=4, end_row=rowCurrent - 1, end_column=7)
        worksheet.merge_cells(start_row=rowCurrent - 1, start_column=8, end_row=rowCurrent - 1, end_column=11)


        for i in range(2, col):
            worksheet.cell(row=rowCurrent - 1, column=i).border = cell_border
            worksheet.cell(row=rowCurrent, column=i).border = cell_border
            worksheet.cell(row=rowCurrent - 1, column=i).font = font_2
            worksheet.cell(row=rowCurrent, column=i).font = font_2
            worksheet.cell(row=rowCurrent, column=i).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=rowCurrent - 1, column=i).alignment = Alignment(horizontal='center', vertical='center')
            
        rowCurrent += 1
        worksheet.cell(row=rowCurrent, column=2).value  = 1
        worksheet.cell(row=rowCurrent, column=3).value  = "-"
        worksheet.cell(row=rowCurrent, column=4).value  = "-"
        worksheet.cell(row=rowCurrent, column=5).value  = "-"
        worksheet.cell(row=rowCurrent, column=6).value  = "-"
        worksheet.cell(row=rowCurrent, column=7).value  = "-"
        worksheet.cell(row=rowCurrent, column=8).value  = "-"
        worksheet.cell(row=rowCurrent, column=9).value  = "-"
        worksheet.cell(row=rowCurrent, column=10).value  = "-"
        worksheet.cell(row=rowCurrent, column=11).value  = "-"

        if count == 1:
            worksheet.cell(row=rowCurrent, column=12).value  = "-"
            worksheet.cell(row=rowCurrent, column=13).value  = "-"

        for i in range(2, col):
            worksheet.cell(row=rowCurrent, column=i).border = cell_border
        rowCurrent += 3
    res = workbook.get_sheet_names()     
    workbook.save(workSpace(module) + fileName)
    workbook.close()
    return res

def renameFile(pathDir, pathSrc):
    os.rename(pathDir, pathSrc)


