# -*- coding: utf-8 -*-
import openpyxl 
from xml.dom import minidom
import sys
import os
import re
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, numbers
import subprocess
import TUT_GenReport
import dataGen

# Function: Function
# Return: Result excel file 
# Content: PASSED/FAILED
def createResult(Msn):
    if Msn.lower() == "generic":
        pathResult = [TUT_GenReport.workSpace(Msn) + Msn + "\\Test_Result_" + Msn + ".trx"]
    else:
        pathResult = [TUT_GenReport.workSpace(Msn) + "E2x\\Test_Result_E2x.trx", 
                      TUT_GenReport.workSpace(Msn) + "U2x\\Test_Result_U2x.trx"]
    fileName = "Result.xlsx"
    
    #font define
    font = Font(name='Calibri',
                     size=13,
                     bold=True,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FF000000')
    cell_border = Border(left=Side(border_style='thin', color='FF000000'),
                     right=Side(border_style='thin', color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000')
    )

    #Create file result
    workbook = openpyxl.Workbook()
    workbook.active

    for f in pathResult:
        device = f.split("\\")[-2]
        
        worksheet = workbook.create_sheet("Result_" + device)
        xmldoc = minidom.parse(f)
        testRun = xmldoc.getElementsByTagName('TestRun')
        ResultSummary = testRun[0].getElementsByTagName('ResultSummary')
        Counters = ResultSummary[0].getElementsByTagName('Counters')

        rowCurrent = 3
        Lab = ["Total TestCases", "Executed", "Not Executed", "Passed", "Failed"]
        res = ["total", "executed", "notExecuted", "passed", "failed"]
        
        # Write Info all TC: Total, Executed, Passed, Failed
        for index in range(3, 8):
            rowCurrent = index
            worksheet.cell(row = rowCurrent, column = 3).value = Lab[index - 3]
            worksheet.cell(row = rowCurrent, column = 3).border = cell_border
            worksheet.cell(row = rowCurrent, column = 3).font = font
            worksheet.cell(row = rowCurrent, column = 4).value = Counters[0].getAttribute(res[index - 3])
            worksheet.cell(row = rowCurrent, column = 4).border = cell_border
  
        rowCurrent +=2 
        
        res_2 =["No", "TestCase Name", "PASSED/FAILED", "Reason"]
        for index in range(2, 6):
            worksheet.cell(row = rowCurrent, column = index).value = res_2[index - 2]
            worksheet.cell(row = rowCurrent, column = index).border = cell_border
            worksheet.cell(row = rowCurrent, column = index).font = font
        rowCurrent += 1
        index = 1

        #Get result PASSED/FAILED of TC and reason
        for testCase in testRun:
            allTCs = testCase.getElementsByTagName('Results')
            UnitTestResult = allTCs[0].getElementsByTagName('UnitTestResult')
            for tc in UnitTestResult:
                worksheet.cell(row = rowCurrent, column = 2).value = index
                worksheet.cell(row = rowCurrent, column = 3).value = tc.getAttribute('testName')
                resultTC = tc.getAttribute('outcome')

                # check result: PASSED/FAILED
                # If result is Failed then fill Message into Reason column
                reason = "-"
                color = PatternFill(start_color='00B050',
                                    end_color='00B050',
                                    fill_type='solid')
                if resultTC == "Failed":
                    reason = tc.getElementsByTagName('Output')[0].getElementsByTagName('ErrorInfo')[0].getElementsByTagName('Message')[0].firstChild.nodeValue
                    #print("reason:" + str(reason))
                    color = PatternFill(start_color='ED100C',
                                        end_color='ED100C',
                                        fill_type='solid')
                worksheet.cell(row = rowCurrent, column = 4).value = resultTC.upper()
                worksheet.cell(row = rowCurrent, column = 5).value = reason
                worksheet.cell(row = rowCurrent, column = 4).fill = color
                worksheet.cell(row = rowCurrent, column = 2).border = cell_border
                worksheet.cell(row = rowCurrent, column = 3).border = cell_border
                worksheet.cell(row = rowCurrent, column = 4).border = cell_border
                worksheet.cell(row = rowCurrent, column = 5).border = cell_border
                index = index + 1
                rowCurrent += 1
    pathOutput = TUT_GenReport.workSpace(Msn) + fileName
    workbook.save(pathOutput)
    workbook.close()
    getRevisonOfSvn(pathOutput, font, cell_border, "tags", Msn, pathOutput)
    adjustWidths(pathOutput)

# Function: Set width for Column
# Return: None 
def adjustWidths(pathFile):
    workbook = openpyxl.load_workbook(str(pathFile))
    workbook.active
    sheetNames = workbook.get_sheet_names() 
    for sheetName in sheetNames:
        if sheetName == "Cover":
            continue
        worksheet = workbook.get_sheet_by_name(sheetName)

        # Start changing width from column C onwards
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name# Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based) 
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            if str(column) == "A" or str(column) == "B":
                adjusted_width = 10
            elif (str(column) == "C" or str(column) == "E") and max_length > 110:
                adjusted_width = 110
            else: 
                adjusted_width = (max_length + 2) 
            worksheet.column_dimensions[column].width = adjusted_width
    workbook.save(pathFile)
    workbook.close()

#Get revison of SVN_Revision sheet
def getRevisonOfSvn(pathFile, font, cell_border, repo, module, output):
    workbook = openpyxl.load_workbook(str(pathFile))
    workbook.active
    worksheet = workbook['Sheet']
    
    rowCurrent = 3
    worksheet.cell(row = rowCurrent, column = 3).value = "Object"
    worksheet.cell(row = rowCurrent, column = 3).border = cell_border
    worksheet.cell(row = rowCurrent, column = 3).font = font
    worksheet.cell(row = rowCurrent, column = 4).value = "Revision"
    worksheet.cell(row = rowCurrent, column = 4).border = cell_border

    # If msn is Generic then inorge element 1st in list
    # If other msn then get more revision of generic
    # Get list path and name file to get revision svn
    if module.lower() == "generic":
        listSVN, listTitle = dataGen.SvnFileGeneric()
    else:
        listSVN, listTitle = dataGen.SvnFile(module)
    
    for index in range(0, len(listSVN)):
        rowCurrent += 1
        
        # Get revision of svn
        info = get_svn_revision(listSVN[index])
        if info is not None:
            match = re.search('\d+', re.search('Last Changed Rev\: \d+', info).group()).group()
            if match:
                worksheet.cell(row = rowCurrent, column = 3).value = listTitle[index]
                worksheet.cell(row = rowCurrent, column = 3).border = cell_border
                worksheet.cell(row = rowCurrent, column = 3).font = font
                worksheet.cell(row = rowCurrent, column = 4).value = int(match)
                worksheet.cell(row = rowCurrent, column = 4).border = cell_border
    worksheet.title = 'SVN_Revision'
    workbook.save(output)
    workbook.close()

# Function: Read info from svn
# Return:   Information of input
# Paramter: link need to check
def get_svn_revision(input_):
    if input_ is not None:
        p = subprocess.Popen(
            "svn info " + input_, stdout=subprocess.PIPE, shell=True)
        info, err = p.communicate()
    return info
